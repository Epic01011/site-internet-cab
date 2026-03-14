#!/usr/bin/env python3
"""
SEO Programmatique Pipeline — Hayot Expertise
==============================================
Convertit le fichier Excel Nexco en seo_state.json et injecte le contenu
dans les fichiers TypeScript data/blog.ts, data/services.ts, data/secteurs.ts.

Usage:
    python seo_pipeline.py generate     # Génère seo_state.json depuis l'Excel
    python seo_pipeline.py validate     # Vérifie cohérence et doublons
    python seo_pipeline.py inject N     # Injecte N pages (mode placeholder)
    python seo_pipeline.py status       # Affiche l'état du pipeline
"""

import json
import re
import sys
import os
from pathlib import Path

# ─────────────────────────────────────────────────────────────
# Configuration
# ─────────────────────────────────────────────────────────────

EXCEL_PATH = os.path.join(os.path.expanduser("~"), "Downloads", "Structure_Nexco_HorsBlog.xlsx")
STATE_FILE = Path(__file__).parent / "seo_state.json"
DATA_DIR = Path(__file__).parent / "data"

BLOG_TS = DATA_DIR / "blog.ts"
SERVICES_TS = DATA_DIR / "services.ts"
SECTEURS_TS = DATA_DIR / "secteurs.ts"

# Mots-clés pour la catégorisation automatique
KEYWORDS_SECTEURS = [
    "expert-comptable", "avocat", "médecin", "architecte", "kinésithérapeute",
    "pharmacien", "dentiste", "infirmier", "profession libérale", "professions libérales",
    "restauration", "hôtellerie", "chr", "btp", "artisan",
    "startup", "freelance", "consultant", "coach", "influenceur",
    "e-commerce", "dropshipping", "amazon", "marketplace",
    "immobilier", "sci", "lmnp", "marchand de biens",
    "médical", "paramédical", "vétérinaire", "ostéopathe",
    "portage salarial", "association", "ong",
]

KEYWORDS_BLOG = [
    "barème", "bareme", "smic", "calculatrice", "simulateur", "grille",
    "taux", "plafond", "calendrier", "date limite", "actualité", "actualite",
    "réforme", "reforme", "loi de finances", "guide", "comment",
    "définition", "definition", "qu'est-ce", "quest-ce", "tout savoir",
    "faq", "questions", "erreurs", "conseils", "astuces",
    "comparatif", "vs", "versus", "différence", "difference",
    "checklist", "modèle", "modele", "template", "exemple",
    "chiffres", "statistiques", "tendances",
]

# Mapping catégorie → icône Lucide
ICON_MAP_SERVICES = {
    "Comptabilité": "Calculator",
    "Fiscalité": "Building2",
    "Juridique": "Scale",
    "Social": "Users",
    "Immobilier": "Home",
    "Création": "Rocket",
    "Digital": "Monitor",
    "Audit": "ClipboardCheck",
    "Conseil": "Lightbulb",
    "Finance": "TrendingUp",
    "default": "FileText",
}

ICON_MAP_SECTEURS = {
    "Immobilier": "Building",
    "Tech": "Zap",
    "Médical": "HeartPulse",
    "Commerce": "ShoppingCart",
    "Restauration": "UtensilsCrossed",
    "BTP": "HardHat",
    "Libéral": "Briefcase",
    "default": "Building",
}


# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def create_slug(title: str) -> str:
    """Génère un slug URL-friendly à partir d'un titre."""
    slug = title.lower().strip()
    # Remplacement des caractères accentués
    replacements = {
        'à': 'a', 'â': 'a', 'ä': 'a', 'æ': 'ae',
        'é': 'e', 'è': 'e', 'ê': 'e', 'ë': 'e',
        'î': 'i', 'ï': 'i',
        'ô': 'o', 'ö': 'o', 'œ': 'oe',
        'ù': 'u', 'û': 'u', 'ü': 'u',
        'ç': 'c', 'ñ': 'n',
        "'": "-", "'": "-", "«": "", "»": "", '"': "", '"': "",
        ":": "", "!": "", "?": "", ",": "", ";": "", ".": "",
        "(": "", ")": "", "[": "", "]": "", "{": "", "}": "",
        "&": "et", "+": "plus", "@": "at", "%": "pourcent",
        "€": "euros", "$": "dollars", "£": "livres",
    }
    for old, new in replacements.items():
        slug = slug.replace(old, new)
    # Remplace tout non-alphanumérique par un tiret
    slug = re.sub(r'[^a-z0-9]+', '-', slug)
    slug = slug.strip('-')
    # Limite la longueur du slug
    if len(slug) > 80:
        slug = slug[:80].rsplit('-', 1)[0]
    return slug


def categorize_page(title: str, description: str) -> str:
    """Détermine si une page appartient à blog, services ou secteurs."""
    text = f"{title} {description}".lower()

    # Score-based approach
    score_blog = sum(1 for kw in KEYWORDS_BLOG if kw in text)
    score_secteurs = sum(1 for kw in KEYWORDS_SECTEURS if kw in text)

    if score_blog > score_secteurs and score_blog >= 2:
        return "blog"
    if score_secteurs > score_blog and score_secteurs >= 1:
        return "secteurs"
    return "services"  # Default: services


def assign_category_service(title: str, description: str) -> str:
    """Assigne une catégorie de service basée sur le contenu."""
    text = f"{title} {description}".lower()
    if any(kw in text for kw in ["comptab", "bilan", "compte annuel", "tenue"]):
        return "Comptabilité"
    if any(kw in text for kw in ["fiscal", "impôt", "is ", "tva", "taxe", "holding"]):
        return "Fiscalité"
    if any(kw in text for kw in ["juridique", "statut", "création", "cession", "droit"]):
        return "Juridique"
    if any(kw in text for kw in ["paie", "social", "dsn", "urssaf", "smic", "salarié"]):
        return "Social"
    if any(kw in text for kw in ["immobilier", "sci", "lmnp", "loyer", "foncier"]):
        return "Immobilier"
    if any(kw in text for kw in ["audit", "commissaire", "certification"]):
        return "Audit"
    if any(kw in text for kw in ["conseil", "stratégie", "accompagnement", "pilotage"]):
        return "Conseil"
    if any(kw in text for kw in ["digital", "facturation électronique", "erp", "logiciel"]):
        return "Digital"
    return "Conseil"


def assign_category_blog(title: str, description: str) -> str:
    """Assigne une catégorie de blog basée sur le contenu."""
    text = f"{title} {description}".lower()
    if any(kw in text for kw in ["entrepreneuriat", "création", "micro", "auto"]):
        return "Entrepreneuriat"
    if any(kw in text for kw in ["fiscal", "impôt", "tva", "is ", "ir ", "holding"]):
        return "Fiscalité"
    if any(kw in text for kw in ["immobilier", "sci", "lmnp", "loyer"]):
        return "Immobilier"
    if any(kw in text for kw in ["actualité", "réforme", "loi", "budget", "smic"]):
        return "Actualités"
    if any(kw in text for kw in ["juridique", "droit", "statut", "tribunal"]):
        return "Juridique"
    return "Fiscalité"


def get_icon_service(category: str) -> str:
    return ICON_MAP_SERVICES.get(category, ICON_MAP_SERVICES["default"])


def get_icon_secteur(title: str) -> str:
    text = title.lower()
    if any(kw in text for kw in ["immobilier", "sci", "lmnp"]):
        return "Building"
    if any(kw in text for kw in ["startup", "tech", "digital", "saas"]):
        return "Zap"
    if any(kw in text for kw in ["médecin", "santé", "médical", "infirmier", "pharmacien"]):
        return "HeartPulse"
    if any(kw in text for kw in ["restaurant", "hôtel", "chr", "brasserie"]):
        return "UtensilsCrossed"
    if any(kw in text for kw in ["avocat", "notaire", "architecte", "libéral"]):
        return "Briefcase"
    if any(kw in text for kw in ["commerce", "e-commerce", "boutique", "magasin"]):
        return "ShoppingCart"
    if any(kw in text for kw in ["btp", "construction", "artisan"]):
        return "HardHat"
    return "Building"


def extract_seo_keywords(title: str, description: str) -> list:
    """Extrait des mots-clés SEO pertinents."""
    keywords = []
    # Ajoute le titre nettoyé
    clean_title = re.sub(r'[:\-–—|]', '', title).strip()
    keywords.append(clean_title)
    # Ajoute des variantes avec "Hayot"
    keywords.append(f"{clean_title} expert-comptable Paris")
    # Extrait les termes clés de la description
    if description and description != "NA":
        # Premier 100 caractères
        short_desc = description[:100].strip()
        if short_desc:
            keywords.append(short_desc)
    return keywords


# ─────────────────────────────────────────────────────────────
# Commands
# ─────────────────────────────────────────────────────────────

def cmd_generate():
    """Génère seo_state.json à partir du fichier Excel."""
    try:
        import openpyxl
    except ImportError:
        print("❌ Module openpyxl requis. Installez-le avec: pip install openpyxl")
        sys.exit(1)

    if not os.path.exists(EXCEL_PATH):
        print(f"❌ Fichier Excel non trouvé: {EXCEL_PATH}")
        sys.exit(1)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    seo_state = {
        "status": "ready",
        "total_tasks": 0,
        "completed_keywords": [],
        "git_status": "clean",
        "pending_tasks": [],
        "stats": {
            "blog": 0,
            "services": 0,
            "secteurs": 0,
        }
    }

    seen_slugs = set()
    skipped = 0

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        url, title, description, page_type = row

        if not title:
            skipped += 1
            continue

        title = str(title).strip()
        description = str(description).strip() if description else "NA"

        # Nettoie les descriptions trop longues ou invalides
        if len(description) > 500:
            description = description[:500].rsplit('.', 1)[0] + '.'

        slug = create_slug(title)

        # Déduplication
        if slug in seen_slugs:
            # Ajoute un suffixe
            i = 2
            while f"{slug}-{i}" in seen_slugs:
                i += 1
            slug = f"{slug}-{i}"
        seen_slugs.add(slug)

        target_file = categorize_page(title, description)
        seo_state["stats"][target_file] += 1

        task = {
            "keyword_slug": slug,
            "competitor_url": str(url).strip() if url else "",
            "competitor_title": title,
            "competitor_desc": description,
            "target_file": target_file,
            "status": "pending",
        }
        seo_state["pending_tasks"].append(task)

    seo_state["total_tasks"] = len(seo_state["pending_tasks"])

    with open(STATE_FILE, 'w', encoding='utf-8') as f:
        json.dump(seo_state, f, ensure_ascii=False, indent=2)

    print(f"✅ {seo_state['total_tasks']} tâches générées dans seo_state.json")
    print(f"   📝 Blog:     {seo_state['stats']['blog']}")
    print(f"   ⚙️  Services: {seo_state['stats']['services']}")
    print(f"   🎯 Secteurs: {seo_state['stats']['secteurs']}")
    if skipped:
        print(f"   ⚠️  Lignes ignorées (vides): {skipped}")


def cmd_validate():
    """Valide la cohérence du seo_state.json et des fichiers TS existants."""
    if not STATE_FILE.exists():
        print("❌ seo_state.json non trouvé. Lancez d'abord: python seo_pipeline.py generate")
        sys.exit(1)

    with open(STATE_FILE, 'r', encoding='utf-8') as f:
        state = json.load(f)

    # Vérification des slugs uniques
    slugs = [t["keyword_slug"] for t in state["pending_tasks"]]
    duplicates = [s for s in set(slugs) if slugs.count(s) > 1]

    # Vérifier les slugs existants dans les fichiers TS
    existing_slugs = set()
    for ts_file in [BLOG_TS, SERVICES_TS, SECTEURS_TS]:
        if ts_file.exists():
            content = ts_file.read_text(encoding='utf-8')
            found = re.findall(r'slug:\s*["\']([^"\']+)["\']', content)
            existing_slugs.update(found)

    conflicts = [s for s in slugs if s in existing_slugs]

    print(f"📊 Validation du pipeline")
    print(f"   Total tâches: {len(state['pending_tasks'])}")
    print(f"   Complétées:   {len(state['completed_keywords'])}")
    print(f"   En attente:   {len([t for t in state['pending_tasks'] if t['status'] == 'pending'])}")

    if duplicates:
        print(f"   ❌ Slugs en double: {duplicates[:10]}...")
    else:
        print(f"   ✅ Aucun slug en double")

    if conflicts:
        print(f"   ⚠️  {len(conflicts)} conflits avec les pages existantes: {conflicts[:5]}...")
    else:
        print(f"   ✅ Aucun conflit avec les pages existantes")


def cmd_inject(batch_size: int = 5):
    """Injecte un batch de pages placeholder dans les fichiers TypeScript."""
    if not STATE_FILE.exists():
        print("❌ seo_state.json non trouvé. Lancez d'abord: python seo_pipeline.py generate")
        sys.exit(1)

    with open(STATE_FILE, 'r', encoding='utf-8') as f:
        state = json.load(f)

    pending = [t for t in state["pending_tasks"] if t["status"] == "pending"]
    if not pending:
        print("✅ Toutes les tâches ont été traitées!")
        return

    batch = pending[:batch_size]
    print(f"📦 Injection de {len(batch)} pages...")

    # Grouper par fichier cible
    by_target = {"blog": [], "services": [], "secteurs": []}
    for task in batch:
        by_target[task["target_file"]].append(task)

    # Injecter dans chaque fichier
    for target, tasks in by_target.items():
        if not tasks:
            continue

        if target == "blog":
            _inject_blog(tasks)
        elif target == "services":
            _inject_services(tasks)
        elif target == "secteurs":
            _inject_secteurs(tasks)

        # Marquer comme complétées
        for task in tasks:
            task["status"] = "completed"
            state["completed_keywords"].append(task["keyword_slug"])

    # Sauvegarder l'état
    with open(STATE_FILE, 'w', encoding='utf-8') as f:
        json.dump(state, f, ensure_ascii=False, indent=2)

    completed = len(state["completed_keywords"])
    total = state["total_tasks"]
    print(f"✅ Batch injecté! Progression: {completed}/{total} ({completed*100//total}%)")
    print(f"   ⚠️  Lancez 'npm run build' pour vérifier le build!")


def _inject_blog(tasks: list):
    """Injecte des entrées BlogPost dans data/blog.ts."""
    content = BLOG_TS.read_text(encoding='utf-8')

    new_entries = []
    for task in tasks:
        category = assign_category_blog(task["competitor_title"], task["competitor_desc"])
        keywords = extract_seo_keywords(task["competitor_title"], task["competitor_desc"])
        desc = task["competitor_desc"] if task["competitor_desc"] != "NA" else f"Guide complet : {task['competitor_title']}. Conseils et accompagnement par Hayot Expertise, expert-comptable Paris 8."

        # Génère le contenu placeholder (sera enrichi par GPT-4o dans le pipeline complet)
        markdown_content = _generate_placeholder_content(task["competitor_title"], desc)

        entry = f"""  {{
    slug: "{task["keyword_slug"]}",
    title: "{_escape_ts(task["competitor_title"])}",
    category: "{category}",
    excerpt: "{_escape_ts(desc[:200])}",
    author: "Samuel HAYOT",
    publishedDate: "2026-03-14",
    content: `
{markdown_content}
    `,
    seoKeywords: [{', '.join(f'"{_escape_ts(kw)}"' for kw in keywords[:5])}],
    relatedPosts: [],
    relatedServices: ["expertise-comptable-paris-8"]
  }}"""
        new_entries.append(entry)

    # Insère avant le dernier '];'
    insert_point = content.rfind('];')
    if insert_point == -1:
        print("❌ Impossible de trouver le point d'insertion dans blog.ts")
        return

    # Ajoute une virgule après le dernier objet existant
    before = content[:insert_point].rstrip()
    if before.endswith('}'):
        before += ','

    new_content = before + '\n' + ',\n'.join(new_entries) + '\n' + content[insert_point:]
    BLOG_TS.write_text(new_content, encoding='utf-8')
    print(f"   📝 blog.ts: +{len(tasks)} entrées")


def _inject_services(tasks: list):
    """Injecte des entrées Service dans data/services.ts."""
    content = SERVICES_TS.read_text(encoding='utf-8')

    new_entries = []
    for task in tasks:
        category = assign_category_service(task["competitor_title"], task["competitor_desc"])
        icon = get_icon_service(category)
        keywords = extract_seo_keywords(task["competitor_title"], task["competitor_desc"])
        desc = task["competitor_desc"] if task["competitor_desc"] != "NA" else f"{task['competitor_title']} — Accompagnement expert par Hayot Expertise, cabinet comptable Paris 8."

        markdown_content = _generate_placeholder_content(task["competitor_title"], desc)

        entry = f"""  {{
    slug: "{task["keyword_slug"]}",
    title: "{_escape_ts(task["competitor_title"])}",
    category: "{category}",
    description: "{_escape_ts(desc[:250])}",
    content: `
{markdown_content}
    `,
    seoKeywords: [{', '.join(f'"{_escape_ts(kw)}"' for kw in keywords[:5])}],
    icon: "{icon}",
    relatedServices: ["expertise-comptable-paris-8"],
    relatedSecteurs: []
  }}"""
        new_entries.append(entry)

    insert_point = content.rfind('];')
    if insert_point == -1:
        print("❌ Impossible de trouver le point d'insertion dans services.ts")
        return

    before = content[:insert_point].rstrip()
    if before.endswith('}'):
        before += ','

    new_content = before + '\n' + ',\n'.join(new_entries) + '\n' + content[insert_point:]
    SERVICES_TS.write_text(new_content, encoding='utf-8')
    print(f"   ⚙️  services.ts: +{len(tasks)} entrées")


def _inject_secteurs(tasks: list):
    """Injecte des entrées Secteur dans data/secteurs.ts."""
    content = SECTEURS_TS.read_text(encoding='utf-8')

    new_entries = []
    for task in tasks:
        icon = get_icon_secteur(task["competitor_title"])
        keywords = extract_seo_keywords(task["competitor_title"], task["competitor_desc"])
        desc = task["competitor_desc"] if task["competitor_desc"] != "NA" else f"Expert-comptable spécialisé : {task['competitor_title']}. Cabinet Hayot Expertise, Paris 8."

        markdown_content = _generate_placeholder_content(task["competitor_title"], desc)

        entry = f"""  {{
    slug: "{task["keyword_slug"]}",
    title: "{_escape_ts(task["competitor_title"])}",
    description: "{_escape_ts(desc[:250])}",
    content: `
{markdown_content}
    `,
    seoKeywords: [{', '.join(f'"{_escape_ts(kw)}"' for kw in keywords[:5])}],
    icon: "{icon}",
    relatedServices: ["expertise-comptable-paris-8"],
    stats: [
      {{ label: "Clients accompagnés", value: "100+" }},
      {{ label: "Satisfaction client", value: "98%" }}
    ]
  }}"""
        new_entries.append(entry)

    insert_point = content.rfind('];')
    if insert_point == -1:
        print("❌ Impossible de trouver le point d'insertion dans secteurs.ts")
        return

    before = content[:insert_point].rstrip()
    if before.endswith('}'):
        before += ','

    new_content = before + '\n' + ',\n'.join(new_entries) + '\n' + content[insert_point:]
    SECTEURS_TS.write_text(new_content, encoding='utf-8')
    print(f"   🎯 secteurs.ts: +{len(tasks)} entrées")


def _generate_placeholder_content(title: str, description: str) -> str:
    """Génère du contenu Markdown placeholder (à remplacer par GPT-4o)."""
    safe_title = title.replace('`', "'")
    safe_desc = description.replace('`', "'").replace('\n', ' ')[:300]

    return f"""# {safe_title}

## Votre expert-comptable vous accompagne

{safe_desc}

### Pourquoi choisir Hayot Expertise ?

Situé au **58 rue de Monceau, Paris 8ème**, notre cabinet d'expertise comptable vous offre un accompagnement personnalisé et des solutions sur-mesure.

**Samuel HAYOT**, expert-comptable diplômé, et son équipe vous accompagnent dans toutes vos démarches avec une approche digitale et humaine.

### Nos points forts

- ✅ **Expertise reconnue** : Plus de 10 ans d'expérience
- ✅ **Outils digitaux** : Pennylane, facturation électronique
- ✅ **Proximité** : Cabinet situé dans le 8ème arrondissement de Paris
- ✅ **Réactivité** : Réponse sous 24h

### Contactez-nous

📍 58 rue de Monceau, 75008 Paris
📞 01 48 85 61 01
📧 contact@hayot-expertise.fr

**Prenez rendez-vous** pour un premier échange gratuit et sans engagement."""


def _escape_ts(text: str) -> str:
    """Échappe les caractères spéciaux pour l'insertion dans du TypeScript."""
    return (text
            .replace('\\', '\\\\')
            .replace('"', '\\"')
            .replace('\n', ' ')
            .replace('\r', '')
            .replace('`', "'"))


def cmd_status():
    """Affiche l'état actuel du pipeline."""
    if not STATE_FILE.exists():
        print("❌ seo_state.json non trouvé. Lancez d'abord: python seo_pipeline.py generate")
        return

    with open(STATE_FILE, 'r', encoding='utf-8') as f:
        state = json.load(f)

    total = state["total_tasks"]
    completed = len(state["completed_keywords"])
    pending = len([t for t in state["pending_tasks"] if t["status"] == "pending"])

    print(f"📊 État du Pipeline SEO Hayot Expertise")
    print(f"   Status:     {state['status']}")
    print(f"   Total:      {total}")
    print(f"   Complétées: {completed} ({completed*100//total if total else 0}%)")
    print(f"   En attente: {pending}")
    print(f"   Répartition:")
    print(f"     📝 Blog:     {state['stats']['blog']}")
    print(f"     ⚙️  Services: {state['stats']['services']}")
    print(f"     🎯 Secteurs: {state['stats']['secteurs']}")

    if completed > 0:
        remaining_hours = pending * 2  # 2h pause entre chaque
        remaining_days = remaining_hours / 24
        print(f"   ⏱️  Temps restant estimé: {remaining_days:.0f} jours (avec pause 2h)")


# ─────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(0)

    command = sys.argv[1].lower()

    if command == "generate":
        cmd_generate()
    elif command == "validate":
        cmd_validate()
    elif command == "inject":
        batch_size = int(sys.argv[2]) if len(sys.argv) > 2 else 5
        cmd_inject(batch_size)
    elif command == "status":
        cmd_status()
    else:
        print(f"❌ Commande inconnue: {command}")
        print(__doc__)
        sys.exit(1)
